import json
from io import BytesIO
from datetime import datetime, timedelta

from django.db.models import Q, Sum, Count, F
from django.db.models.functions import TruncMonth
from django.http import JsonResponse
from django.urls import reverse
from django.utils import timezone
from django.utils.safestring import mark_safe

from enterprise.models import Cashier, Bill, Payment

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from enterprise.models import Installments, PaymentMethod, NFSe
from students.models import MonthlyFee, Student, Payment


def get_context_cashier_data():
        last_cashier = Cashier.objects.order_by('-created_at').first()
        if last_cashier and last_cashier.status == 'closed':
            context = {
                'status': last_cashier.get_status_display(),
                'income_pix': last_cashier.income_pix,
                'income_credit': last_cashier.income_credit,
                'income_debit': last_cashier.income_debit,
                'income_cash': last_cashier.income_cash,
                'total_incomes': last_cashier.total_incomes,
                'bills': last_cashier.bills.all(),
                'expense_pix': last_cashier.expense_pix,
                'expense_boleto': last_cashier.expense_boleto,
                'expense_automatic': last_cashier.expense_automatic,
                'expense_others': last_cashier.expense_others,
                'expense_withdrawal': last_cashier.expense_withdrawal,
                'total_expenses': last_cashier.total_expenses,
                'payments': last_cashier.payments.all(),
                'date_start': last_cashier.created_at,
                'date_end': last_cashier.date_closing,
                'opening_balance': last_cashier.opening_balance,
                'closing_balance': last_cashier.closing_balance
            }
            return context
        start = last_cashier.created_at if last_cashier else datetime.now()
        end = datetime.now()
        payments = Payment.objects.filter(Q(created_at__gte=(start))
                                          | Q(created_at__lte=(end)),
                                          cashier__isnull=True)
        total = payments.aggregate(
            income_pix=Sum('value', filter=Q(payment_method__iexact="pix")),
            income_credit=Sum('value', filter=Q(
                payment_method__iexact="credito")),
            income_debit=Sum('value', filter=Q(
                payment_method__iexact="debito")),
            income_cash=Sum('value', filter=Q(
                payment_method__iexact="dinheiro")),
            total_incomes=Sum('value')
        )
        bill = Bill.objects.select_related(
            'payment_method', 'status').filter(Q(date_payment__gte=(start)) |
                                               Q(date_payment__lte=(end)), cashier__isnull=True)
        pay = bill.filter(Q(status__status__iexact='pago') | Q(payment_method__method__icontains='automatico')).aggregate(
            total_expenses=Sum('value'),
            expense_pix=Sum('value', filter=Q(
                payment_method__method__iexact='pix')),
            expense_boleto=Sum('value', filter=Q(
                payment_method__method__iexact='boleto')),
            expense_automatic=Sum('value', filter=Q(
                payment_method__method__iexact='deb. automatico')),
            expense_others=Sum('value', filter=Q(payment_method__method__iexact='credito') | Q(
                payment_method__method__iexact='debito')),
        )
        income = total['total_incomes'] or 0
        expense = pay['total_expenses'] or 0
        last_balance = Cashier.objects.filter(
            status='closed').order_by('-created_at').first()
        closing_balance = (last_balance.closing_balance + income - expense) if last_balance else (income - expense)
        expense_withdrawal = last_cashier.expense_withdrawal if last_cashier and last_cashier.status == 'open' else 0
        context = {
            'status': last_cashier.get_status_display() if last_cashier else "Fechado",
            'income_pix': total['income_pix'] if total['income_pix'] else 0,
            'income_credit': total['income_credit'] if total['income_credit'] else 0,
            'income_debit': total['income_debit'] if total['income_debit'] else 0,
            'income_cash': total['income_cash'] if total['income_cash'] else 0,
            'total_incomes': total['total_incomes'] if total['total_incomes'] else 0,
            'bills': bill,
            'expense_pix': pay['expense_pix'] if pay['expense_pix'] else 0,
            'expense_boleto': pay['expense_boleto'] if pay['expense_boleto'] else 0,
            'expense_automatic': pay['expense_automatic'] if pay['expense_automatic'] else 0,
            'expense_others': pay['expense_others'] if pay['expense_others'] else 0,
            'expense_withdrawal': expense_withdrawal,
            'total_expenses': pay['total_expenses'] if pay['total_expenses'] else 0,
            'payments': payments,
            'date_start': start,
            'opening_balance': last_balance.closing_balance if last_balance else 0,
            'closing_balance': closing_balance if closing_balance else 0
        }
        return context
    
def create_new_register_cashier():
    try:
        if Cashier.objects.filter(status='open').exists():
            return JsonResponse({'status': 'warning', 'title': 'Abertura de Caixa', 'message': 'Já existe um caixa aberto!'}, status=400)
        last_cashier = Cashier.objects.order_by('-created_at').first()
        new_cashier = Cashier.objects.create(
            opening_balance=last_cashier.closing_balance if last_cashier else 0,
            status='open',
        )
        return JsonResponse({'status': 'success', 'title': 'Abertura de Caixa', 'message': 'Caixa aberto com sucesso!'}, status=201)
    except Exception as e:
        return JsonResponse({'status': 'error', 'title': 'Erro ao abrir caixa', 'message': f'{e}'}, status=400)
    
    
def close_cashier(context, withdrawalValue, closing_balance):
    try:
        last_cashier = Cashier.objects.order_by('-created_at').first()
        last_cashier.status = 'closed'
        last_cashier.date_closing = datetime.now()
        last_cashier.total_incomes = context['total_incomes']
        last_cashier.total_expenses = context['total_expenses'] + withdrawalValue
        last_cashier.closing_balance = closing_balance
        last_cashier.income_pix = context['income_pix']
        last_cashier.income_credit = context['income_credit']
        last_cashier.income_debit = context['income_debit']
        last_cashier.income_cash = context['income_cash']
        last_cashier.expense_pix = context['expense_pix']
        last_cashier.expense_boleto = context['expense_boleto']
        last_cashier.expense_automatic = context['expense_automatic']
        last_cashier.expense_others = context['expense_others']
        last_cashier.expense_withdrawal = withdrawalValue
        last_cashier.save()
        # Atualizar os pagamentos e contas vinculando ao caixa fechado
        context['bills'].update(cashier=last_cashier)
        context['payments'].update(cashier=last_cashier)
        return JsonResponse({'status': 'success', 'title': 'Fechamento do Caixa', 'message': f'Caixa fechado com sucesso! Saldo final R${closing_balance}, Retirado: R${withdrawalValue}'}, status=200)
    except Exception as e:
        return JsonResponse({'status': 'error', 'title': 'Erro ao fechar Caixa', 'message': str(e)}, status=400)


def create_file_xlsx_cashier(cashier):
    wb = Workbook()
    ws = wb.active
    ws.title = "Fechamento de Caixa"
    ws.sheet_view.showGridLines = False
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="d9d9d9")
    green_fill = PatternFill("solid", fgColor="d8e4bc")
    red_fill = PatternFill("solid", fgColor="f4cccc")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

      # === CABEÇALHO ===
    ws.merge_cells("A1:B1")
    ws["A1"] = "Fechamento de Caixa"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = header_fill

    ws["A2"], ws["B2"] = "Situação", cashier.get_status_display()
    ws["A3"], ws["B3"] = "Data de abertura", cashier.created_at.strftime("%d/%m/%Y")
    ws["D2"], ws["E2"] = "Data de fechamento", cashier.date_closing.strftime("%d/%m/%Y") if cashier.date_closing else ""

        # === ENTRADAS ===
    ws["A5"] = "Entradas"
    ws["A5"].font = Font(bold=True, color="006100")
    entries = [
            ("Valor total", cashier.total_incomes or 0),
            ("Dinheiro", cashier.income_cash or 0),
            ("Crédito", cashier.income_credit or 0),
            ("Débito", cashier.income_debit or 0),
            ("Pix", cashier.income_pix or 0),
        ]
    for i, (label, val) in enumerate(entries, start=6):
            ws[f"A{i}"], ws[f"B{i}"] = label, val
            ws[f"B{i}"].number_format = "R$ #,##0.00"

        # === SAÍDAS ===
    ws["A12"] = "Saídas"
    ws["A12"].font = Font(bold=True, color="9c0006")
    ws["A13"], ws["B13"] = "Valor total", cashier.total_expenses or 0
    ws["B13"].number_format = "R$ #,##0.00"

        # === SALDO ===
    ws["A15"] = "Saldo total"
    ws["A15"].font = bold
    ws["B15"] = (cashier.total_incomes or 0) - (cashier.total_expenses or 0)
    ws["B15"].number_format = "R$ #,##0.00"
    ws["B15"].font = bold

        # === MOVIMENTOS ===
    start_row = 17
    ws[f"A{start_row}"] = "Movimentos"
    ws[f"A{start_row}"].font = Font(bold=True, color="1f4e78")

    headers = ["Data", "Descrição", "Origem", "Método de pagamento", "Valor", "Tipo"]
    for col, header in enumerate(headers, start=1):
        c = ws.cell(row=start_row + 1, column=col, value=header)
        c.font = bold
        c.alignment = Alignment(horizontal="center")
        c.fill = header_fill
        c.border = thin_border

    row = start_row + 2

        # === PAGAMENTOS (Entradas) ===
    for p in cashier.payments.select_related("montlhyfee"):
        ws.cell(row=row, column=1, value=p.created_at.strftime("%d/%m/%Y %H:%M"))
        ws.cell(row=row, column=2, value=getattr(p.montlhyfee, "student_name", ""))
        ws.cell(row=row, column=3, value="Contas a receber")
        ws.cell(row=row, column=4, value=p.payment_method)
        ws.cell(row=row, column=5, value=p.value).number_format = "R$ #,##0.00"
        ws.cell(row=row, column=6, value="Entrada")
        row += 1

        # === BILLS (Saídas) ===
    for b in cashier.bills.select_related("payment_method"):
            ws.cell(row=row, column=1, value=b.date_payment.strftime("%d/%m/%Y") if b.date_payment else "")
            ws.cell(row=row, column=2, value=b.description)
            ws.cell(row=row, column=3, value="Contas a pagar")
            ws.cell(row=row, column=4, value=b.payment_method.method if b.payment_method else "")
            ws.cell(row=row, column=5, value=b.value).number_format = "R$ #,##0.00"
            ws.cell(row=row, column=6, value="Saída")
            row += 1

        # === FORMATAÇÃO FINAL ===
    for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Bordas
    for r in range(start_row + 2, row):
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).border = thin_border
                ws.cell(row=r, column=c).alignment = Alignment(vertical="center")

        # === RETORNO ===
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    file_name = f"Fechamento_de_Caixa_{cashier.date_closing:%d-%m-%Y}.xlsx"
    
    return (buffer, file_name)



def get_context_homeview():
    today = datetime.now()
    actives_students = Student.objects.filter(
            status__status__iexact='Ativo')
    end_date = today+timedelta(days=10)
    monthly_fees_due = MonthlyFee.objects.filter(
            due_date__range=(today, end_date), paid=False)
    date_start = today - timedelta(days=10)
    date_end = today - timedelta(days=1)
    monthly_fees_overdue = MonthlyFee.objects.filter(
            due_date__range=(date_start, date_end), paid=False)
    installments = Installments.objects.all()
    bill_events = (
            Bill.objects.annotate(
                event_date=F('due_date'))
            .values('event_date')
            .annotate(count=Count('id'))
            .order_by('event_date')
        )
    calendar_events = [
            {
                "date": event.get('event_date').isoformat(),
                "count": event.get('count', 0),
            }
            for event in bill_events
            if event.get('event_date') is not None
        ]
    context = {
            'actives_total': actives_students.count(),
            'actives_students': actives_students,
            'monthly_fees_due_total': monthly_fees_due.count(),
            'monthly_fees_due': monthly_fees_due,
            'monthly_fees_overdue_total': monthly_fees_overdue.count(),
            'monthly_fees_overdue': monthly_fees_overdue,
            'today': today,
            'payment_methods':  PaymentMethod.objects.filter(
                applies_to__icontains='students'),
            'calendar_events': mark_safe(json.dumps(calendar_events)),
            'accounts_url': reverse('list_bill'),
            'students_active_url': f"{reverse('list_student')}?filter=ativo",
            'installments': installments,
        }
    return context


def get_dashboard_context():
    today = timezone.localdate()
    next_week = today + timedelta(days=7)
    six_months_ago = today - timedelta(days=180)

    students_qs = Student.objects.select_related('status', 'plan')
    total_students = students_qs.count()
    active_students = students_qs.filter(status__status__iexact='Ativo').count()
    inactive_students = students_qs.filter(status__status__iexact='Inativo').count()

    monthlyfees_qs = MonthlyFee.objects.select_related('student', 'plan')
    monthly_fee_values = monthlyfees_qs.aggregate(
        total_value=Sum('amount'),
        paid_value=Sum('amount', filter=Q(paid=True)),
        pending_value=Sum('amount', filter=Q(paid=False)),
    )
    monthly_fee_values = {
        key: value or 0 for key, value in monthly_fee_values.items()
    }

    monthly_fee_counts = monthlyfees_qs.aggregate(
        total=Count('id'),
        paid_count=Count('id', filter=Q(paid=True)),
        pending_count=Count('id', filter=Q(paid=False)),
        overdue_count=Count('id', filter=Q(paid=False, due_date__lt=today)),
        due_next_count=Count(
            'id',
            filter=Q(paid=False, due_date__range=(today, next_week))
        ),
    )
    monthly_fee_counts = {
        key: value or 0 for key, value in monthly_fee_counts.items()
    }

    upcoming_fees_qs = monthlyfees_qs.filter(
        paid=False, due_date__range=(today, next_week)
    ).order_by('due_date')[:6]
    overdue_fees_qs = monthlyfees_qs.filter(
        paid=False, due_date__lt=today
    ).order_by('due_date')[:6]

    def serialize_fee(fee):
        return {
            'name': fee.student.name if fee.student else fee.student_name or 'Sem identificação',
            'plan': fee.plan.name_plan if fee.plan else 'Sem plano',
            'due_date': fee.due_date,
            'amount': fee.amount if fee.amount is not None else 0,
        }

    upcoming_fees = [serialize_fee(fee) for fee in upcoming_fees_qs]
    overdue_fees = [serialize_fee(fee) for fee in overdue_fees_qs]

    bills_qs = Bill.objects.select_related('status', 'payment_method')
    bills_summary = {
        'open_count': bills_qs.filter(date_payment__isnull=True).count(),
        'paid_count': bills_qs.filter(date_payment__isnull=False).count(),
        'overdue_count': bills_qs.filter(
            date_payment__isnull=True, due_date__lt=today
        ).count(),
        'total_value': bills_qs.aggregate(total=Sum('value'))['total'] or 0,
    }
    bills_by_status = list(
        bills_qs.values('status__status')
        .annotate(total=Count('id'), total_value=Sum('value'))
        .order_by('-total')
    )
    for bill in bills_by_status:
        bill['status'] = bill.pop('status__status') or 'Sem status'
        bill['total_value'] = bill['total_value'] or 0

    payment_method_stats = list(
        Payment.objects.values('payment_method')
        .annotate(total=Sum('value'), count=Count('id'))
        .order_by('-total')
    )
    for stat in payment_method_stats:
        stat['method'] = stat.pop('payment_method') or 'Não informado'
        stat['total'] = float(stat['total'] or 0)
        stat['count'] = stat['count'] or 0

    payment_method_labels = [stat['method']
                             for stat in payment_method_stats]
    payment_method_totals = [stat['total']
                             for stat in payment_method_stats]

    monthly_cashflow_qs = (
        monthlyfees_qs.filter(due_date__gte=six_months_ago)
        .annotate(month=TruncMonth('due_date'))
        .values('month')
        .annotate(
            billed=Sum('amount'),
            received=Sum('amount', filter=Q(paid=True)),
        )
        .order_by('month')
    )
    monthly_cashflow_labels = []
    monthly_cashflow_billed = []
    monthly_cashflow_received = []
    for entry in monthly_cashflow_qs:
        month = entry['month']
        if month:
            monthly_cashflow_labels.append(month.strftime('%b/%y'))
        else:
            monthly_cashflow_labels.append('Sem data')
        monthly_cashflow_billed.append(float(entry['billed'] or 0))
        monthly_cashflow_received.append(float(entry['received'] or 0))

    students_by_plan = list(
        students_qs.values('plan__name_plan')
        .annotate(count=Count('id'))
        .order_by('-count')[:5]
    )
    for plan in students_by_plan:
        plan['name'] = plan.pop('plan__name_plan') or 'Sem plano'
        plan['percentage'] = round(
            (plan['count'] / total_students) * 100, 1
        ) if total_students else 0

    nfse_qs = NFSe.objects.select_related('student')
    nfse_summary = {
        'total': nfse_qs.count(),
        'this_month': nfse_qs.filter(
            issue_date__year=today.year, issue_date__month=today.month
        ).count(),
    }
    recent_invoices = nfse_qs.order_by('-issue_date')[:5]

    context = {
        'students_summary': {
            'total': total_students,
            'active': active_students,
            'inactive': inactive_students,
            'active_percentage': round(
                (active_students / total_students) * 100, 1
            ) if total_students else 0,
        },
        'monthly_fee_values': monthly_fee_values,
        'monthly_fee_counts': monthly_fee_counts,
        'monthly_fee_distribution': {
            'paid_pct': round(
                (monthly_fee_counts['paid_count'] / monthly_fee_counts['total']) * 100, 1
            ) if monthly_fee_counts['total'] else 0,
            'pending_pct': round(
                (monthly_fee_counts['pending_count'] / monthly_fee_counts['total']) * 100, 1
            ) if monthly_fee_counts['total'] else 0,
            'overdue_pct': round(
                (monthly_fee_counts['overdue_count'] / monthly_fee_counts['total']) * 100, 1
            ) if monthly_fee_counts['total'] else 0,
        },
        'upcoming_fees': upcoming_fees,
        'overdue_fees': overdue_fees,
        'bills_summary': bills_summary,
        'bills_by_status': bills_by_status,
        'payment_method_stats': payment_method_stats,
        'payment_method_labels': mark_safe(
            json.dumps(payment_method_labels)
        ),
        'payment_method_totals': mark_safe(
            json.dumps(payment_method_totals)
        ),
        'monthly_cashflow_labels': mark_safe(
            json.dumps(monthly_cashflow_labels)
        ),
        'monthly_cashflow_billed': mark_safe(
            json.dumps(monthly_cashflow_billed)
        ),
        'monthly_cashflow_received': mark_safe(
            json.dumps(monthly_cashflow_received)
        ),
        'students_by_plan': students_by_plan,
        'nfse_summary': nfse_summary,
        'recent_invoices': recent_invoices,
    }
    return context
