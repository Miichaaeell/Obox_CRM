import json

from io import BytesIO
from datetime import datetime, timedelta

from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Count, F, Q, Sum
from django.shortcuts import render
from django.urls import reverse, reverse_lazy
from django.utils.safestring import mark_safe
from django.views.generic import CreateView, ListView, TemplateView, UpdateView, View
from django.http import JsonResponse, FileResponse

from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.views import APIView

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from core. functions import get_context_cashier_data, create_new_register_cashier, close_cashier
from enterprise.forms import PaymentMethodForm, PlanForm
from enterprise.models import Bill, PaymentMethod, Plan, StatusBill, Cashier, Installments
from enterprise.serializers import (
    BillSerializer,
    NFESerializer,
)
from students.models import MonthlyFee, Student, Payment




class EnterpriseHomeView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        today = datetime.now()
        actives_students = Student.objects.filter(
            status__status__iexact='Ativo')
        end_date = today+timedelta(days=10)
        monthly_fees_due = MonthlyFee.objects.filter(
            due_date__range=(today, end_date), paid=False)
        date_start = today - timedelta(days=10)
        date_end = today - timedelta(days=1)
        monthly_fees_overdue = MonthlyFee.objects.filter(
            due_date__range=(date_start, date_end), paid=False)
        installments = Installments.objects.all()
        bill_events = (
            Bill.objects.annotate(
                event_date=F('due_date'))
            .values('event_date')
            .annotate(count=Count('id'))
            .order_by('event_date')
        )
        calendar_events = [
            {
                "date": event.get('event_date').isoformat(),
                "count": event.get('count', 0),
            }
            for event in bill_events
            if event.get('event_date') is not None
        ]
        context = {
            'actives_total': actives_students.count(),
            'actives_students': actives_students,
            'monthly_fees_due_total': monthly_fees_due.count(),
            'monthly_fees_due': monthly_fees_due,
            'monthly_fees_overdue_total': monthly_fees_overdue.count(),
            'monthly_fees_overdue': monthly_fees_overdue,
            'today': today,
            'payment_methods':  PaymentMethod.objects.filter(
                applies_to__icontains='students'),
            'calendar_events': mark_safe(json.dumps(calendar_events)),
            'accounts_url': reverse('list_bill'),
            'students_active_url': f"{reverse('list_student')}?filter=ativo",
            'installments': installments,
        }
        return render(request, 'home.html', context)


class FlowCashierView(LoginRequiredMixin, TemplateView):
    template_name = 'flow_cashier.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cashiers = Cashier.objects.all().order_by('-created_at')
        context['cashiers'] = cashiers
        context['url_download'] = reverse('download_cashier')
        
        return context




class DownloadCashierFlowView(LoginRequiredMixin, View):
    """Gera relatório do fechamento de caixa em uma única aba formatada."""

    def get(self, request, *args, **kwargs):
        cashier_id = request.GET.get("pk")
        try:
            cashier = Cashier.objects.get(id=cashier_id)
        except Cashier.DoesNotExist:
            return JsonResponse({
                "status": "error",
                "message": "Caixa não encontrado."
            }, status=404)

        # === CONFIGURAÇÕES DO EXCEL ===
        wb = Workbook()
        ws = wb.active
        ws.title = "Fechamento de Caixa"
        ws.sheet_view.showGridLines = False

        bold = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="d9d9d9")
        green_fill = PatternFill("solid", fgColor="d8e4bc")
        red_fill = PatternFill("solid", fgColor="f4cccc")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))

        # === CABEÇALHO ===
        ws.merge_cells("A1:B1")
        ws["A1"] = "Fechamento de Caixa"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].fill = header_fill

        ws["A2"], ws["B2"] = "Situação", cashier.get_status_display()
        ws["A3"], ws["B3"] = "Data de abertura", cashier.created_at.strftime("%d/%m/%Y")
        ws["D2"], ws["E2"] = "Data de fechamento", cashier.date_closing.strftime("%d/%m/%Y") if cashier.date_closing else ""

        # === ENTRADAS ===
        ws["A5"] = "Entradas"
        ws["A5"].font = Font(bold=True, color="006100")
        entries = [
            ("Valor total", cashier.total_incomes or 0),
            ("Dinheiro", cashier.income_cash or 0),
            ("Crédito", cashier.income_credit or 0),
            ("Débito", cashier.income_debit or 0),
            ("Pix", cashier.income_pix or 0),
        ]
        for i, (label, val) in enumerate(entries, start=6):
            ws[f"A{i}"], ws[f"B{i}"] = label, val
            ws[f"B{i}"].number_format = "R$ #,##0.00"

        # === SAÍDAS ===
        ws["A12"] = "Saídas"
        ws["A12"].font = Font(bold=True, color="9c0006")
        ws["A13"], ws["B13"] = "Valor total", cashier.total_expenses or 0
        ws["B13"].number_format = "R$ #,##0.00"

        # === SALDO ===
        ws["A15"] = "Saldo total"
        ws["A15"].font = bold
        ws["B15"] = (cashier.total_incomes or 0) - (cashier.total_expenses or 0)
        ws["B15"].number_format = "R$ #,##0.00"
        ws["B15"].font = bold

        # === MOVIMENTOS ===
        start_row = 17
        ws[f"A{start_row}"] = "Movimentos"
        ws[f"A{start_row}"].font = Font(bold=True, color="1f4e78")

        headers = ["Data", "Descrição", "Origem", "Método de pagamento", "Valor", "Tipo"]
        for col, header in enumerate(headers, start=1):
            c = ws.cell(row=start_row + 1, column=col, value=header)
            c.font = bold
            c.alignment = Alignment(horizontal="center")
            c.fill = header_fill
            c.border = thin_border

        row = start_row + 2

        # === PAGAMENTOS (Entradas) ===
        for p in cashier.payments.select_related("montlhyfee"):
            ws.cell(row=row, column=1, value=p.created_at.strftime("%d/%m/%Y %H:%M"))
            ws.cell(row=row, column=2, value=getattr(p.montlhyfee, "student_name", ""))
            ws.cell(row=row, column=3, value="Contas a receber")
            ws.cell(row=row, column=4, value=p.payment_method)
            ws.cell(row=row, column=5, value=p.value).number_format = "R$ #,##0.00"
            ws.cell(row=row, column=6, value="Entrada")
            row += 1

        # === BILLS (Saídas) ===
        for b in cashier.bills.select_related("payment_method"):
            ws.cell(row=row, column=1, value=b.date_payment.strftime("%d/%m/%Y") if b.date_payment else "")
            ws.cell(row=row, column=2, value=b.description)
            ws.cell(row=row, column=3, value="Contas a pagar")
            ws.cell(row=row, column=4, value=b.payment_method.method if b.payment_method else "")
            ws.cell(row=row, column=5, value=b.value).number_format = "R$ #,##0.00"
            ws.cell(row=row, column=6, value="Saída")
            row += 1

        # === FORMATAÇÃO FINAL ===
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Bordas
        for r in range(start_row + 2, row):
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).border = thin_border
                ws.cell(row=r, column=c).alignment = Alignment(vertical="center")

        # === RETORNO ===
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        file_name = f"Fechamento_de_Caixa_{cashier.date_closing:%d-%m-%Y}.xlsx"
        return FileResponse(buffer, as_attachment=True, filename=file_name, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


class EnterpriseCashierView(LoginRequiredMixin, View):

    def get(self, request, *args, **kwargs):
        context = get_context_cashier_data()
        return render(request, 'cashier.html', context)

    def post(self, request, *args, **kwargs):
        data = json.loads(request.body)
        action = data.get('action', 'create')
        if action == 'update':
            context = get_context_cashier_data()
            withdrawalValue = int(data.get('withdrawalValue', 0))
            closing_balance = int(data.get('closing_balance', 0))
            print(f'{withdrawalValue} desconto e fechamento {closing_balance}')
            response = close_cashier(context, withdrawalValue, closing_balance)
            return response
        elif action == 'create':
            response = create_new_register_cashier()
            return response


class PaymentMethodListView(LoginRequiredMixin, ListView):
    model = PaymentMethod
    template_name = 'components/_list.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Métodos de Pagamento'
        context['sufix_url'] = 'payment_method'
        return context


class PaymentMethodCreateView(LoginRequiredMixin, CreateView):
    model = PaymentMethod
    form_class = PaymentMethodForm
    template_name = 'components/_create_update.html'
    success_url = reverse_lazy('list_payment_method')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Método de Pagamento'
        context['sufix_url'] = 'payment_method'
        return context


class PaymentMethodUpdateView(LoginRequiredMixin, UpdateView):
    model = PaymentMethod
    form_class = PaymentMethodForm
    template_name = 'components/_create_update.html'
    success_url = reverse_lazy('list_payment_method')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Método de Pagamento'
        context['sufix_url'] = 'payment_method'
        return context


class PlanListView(LoginRequiredMixin, ListView):
    model = Plan
    template_name = 'components/_list.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Planos'
        context['sufix_url'] = 'plan'
        return context


class PlanCreateView(LoginRequiredMixin, CreateView):
    model = Plan
    form_class = PlanForm
    template_name = 'components/_create_update.html'
    success_url = reverse_lazy('list_plan')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Plano'
        context['sufix_url'] = 'plan'
        return context


class PlanUpdateView(LoginRequiredMixin, UpdateView):
    model = Plan
    form_class = PlanForm
    template_name = 'components/_create_update.html'
    success_url = reverse_lazy('list_plan')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Plano'
        context['sufix_url'] = 'plan'
        return context

    # Views for Bill


class BillListView(LoginRequiredMixin, ListView):
    model = Bill
    template_name = 'list_bill.html'
    # paginate_by = 8

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_bill'] = StatusBill.objects.all()
        context['payment_methods'] = PaymentMethod.objects.all()
        context['current_created_date'] = self.request.GET.get(
            'created_date', '')
        return context

    def get_queryset(self):
        queryset = super().get_queryset()
        due_date = self.request.GET.get('due_date')
        search = self.request.GET.get("search", "").strip()

        filters = Q()

        if search:
            # Tentativa 1: usuário digitou no formato dd/mm/yyyy
            try:
                parsed_date = datetime.strptime(search, "%d/%m/%Y").date()
                filters |= Q(due_date=parsed_date)
            except ValueError:
                pass

            # Tentativa 2: usuário digitou no formato yyyy-mm-dd
            try:
                parsed_date = datetime.strptime(search, "%Y-%m-%d").date()
                filters |= Q(due_date=parsed_date)
            except ValueError:
                pass
            try:
                parsed_month = datetime.strptime(search, "%m/%Y")
                filters |= Q(due_date__year=parsed_month.year,
                             due_date__month=parsed_month.month)
            except ValueError:
                pass

            try:
                parsed_month = datetime.strptime(search, "%m/%y")
                filters |= Q(due_date__year=parsed_month.year,
                             due_date__month=parsed_month.month)
            except ValueError:
                pass
            # Caso contrário, trate o search como texto (ex: nome, plano etc.)
            filters |= Q(status__status__icontains=search)

        queryset = Bill.objects.filter(filters)
        if due_date:
            try:
                parsed_date = datetime.strptime(
                    due_date, "%Y-%m-%d").date()
                queryset = queryset.filter(due_date=parsed_date)
            except ValueError:
                pass
        return queryset

# Views NFE's


class NFESListView(LoginRequiredMixin, TemplateView):
    template_name = 'nfes.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        monthlyfees = MonthlyFee.objects.select_related(
            'student', 'plan').filter(paid=True)
        context['monthlyfees'] = monthlyfees
        reference_months = (
            MonthlyFee.objects.order_by('-reference_month')
            .values_list('reference_month', flat=True)
            .distinct()
        )
        context['reference_months'] = list(reference_months)
        return context


# API'S VIEW
class BillDetailAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Bill.objects.all()
    serializer_class = BillSerializer


class BillCreateAPIView(generics.ListCreateAPIView):
    queryset = Bill.objects.all()
    serializer_class = BillSerializer


class NFEAPIView(APIView):
    def post(self, request, *args, **kwargs):

        serializer = NFESerializer(data=request.data)

        if serializer.is_valid():
            data = serializer.validated_data
            students = data['student']
            description = data['description']
            reference_month = data['reference_month']

            # Criar função para rodar em segundo plano e emitir a nota
            for student in students:
                print(
                    f'emitindo nota para {student['id']}, descrição {description},mes de referencia {reference_month} ')

            return Response({
                'status': 'ok',
                'mensagem': f'{len(students)} notas agendadas para emissão.'
            },
                status=status.HTTP_202_ACCEPTED
            )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
